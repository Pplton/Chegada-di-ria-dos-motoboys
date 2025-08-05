import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook

# Criar pasta 'planilhas' se não existir
pasta_planilhas = "planilhas"
os.makedirs(pasta_planilhas, exist_ok=True)

# Caminho esperado do arquivo original
xls_path = os.path.join(pasta_planilhas, "Pedidos.xls")
xlsx_path = os.path.join(pasta_planilhas, "Pedidos_convertido.xlsx")

# Verifica se o arquivo existe
if not os.path.isfile(xls_path):
    print(f"⚠️ O arquivo 'Pedidos.xls' não foi encontrado na pasta '{pasta_planilhas}'.")
    print("➡️ Coloque o arquivo lá e execute o script novamente.")
    input("Pressione Enter para sair...")
    exit()

# Abrir como DataFrame e salvar como .xlsx
df_xls = pd.read_excel(xls_path)
df_xls.to_excel(xlsx_path, index=False)

# Desmesclar células no novo arquivo .xlsx
wb = load_workbook(xlsx_path)
ws = wb.active

for merge in list(ws.merged_cells.ranges):
    min_col, min_row, max_col, max_row = range_boundaries(str(merge))
    cell_value = ws.cell(row=min_row, column=min_col).value
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = cell_value
    ws.unmerge_cells(str(merge))

# Remover colunas A, C e F (ordem decrescente)
ws.delete_cols(6)  # F
ws.delete_cols(3)  # C
ws.delete_cols(1)  # A

# Remover as linhas 1 a 6
ws.delete_rows(1, 6)

wb.save(xlsx_path)

# Agora, use a planilha desmesclada
df = pd.read_excel(xlsx_path)

# Função para determinar o turno com base na hora
def determinar_turno(horario_str):
    try:
        if pd.isna(horario_str) or horario_str == 'Hora inválida':
            return 'indefinido'
        hora = int(horario_str.split(":")[0])
        return 'turno1' if hora < 16 else 'turno2'
    except:
        return 'indefinido'

# Função para formatar datas, tratando inválidas
def formatar_data(data):
    try:
        return pd.to_datetime(data, dayfirst=True, errors='coerce')
    except:
        return pd.NaT

# Criar pasta 'planilhas' se não existir
pasta_planilhas = "planilhas"
os.makedirs(pasta_planilhas, exist_ok=True)

# Caminho esperado do arquivo
file_path = os.path.join(pasta_planilhas, "Pedidos.xls")

# Verifica se o arquivo existe
if not os.path.isfile(file_path):
    print(f"⚠️ O arquivo 'Pedidos.xls' não foi encontrado na pasta '{pasta_planilhas}'.")
    print("➡️ Coloque o arquivo lá e execute o script novamente.")
    input("Pressione Enter para sair...")
    exit()

# Carregar a planilha convertida e desmesclada
df = pd.read_excel(xlsx_path)

# Extrair colunas por índice com tratamento
col_data_agendada = df.iloc[:, 23].apply(formatar_data)  # Coluna W
col_horario_chegada = df.iloc[:, 27].apply(formatar_data)  # Coluna AA
col_empresa = df.iloc[:, 3].fillna('Empresa não informada')  # Coluna D
col_entregador = df.iloc[:, 13].fillna('Entregador não informado')  # Coluna N

# Criar DataFrame tratado
df_resumido = pd.DataFrame()
df_resumido["Data"] = col_data_agendada.dt.strftime("%d/%m/%Y").replace('NaT', 'Data inválida')
df_resumido["Horario agendado"] = col_data_agendada.dt.strftime("%H:%M:%S").replace('NaT', 'Hora inválida')
df_resumido["Horario real de chegada"] = col_horario_chegada.dt.strftime("%H:%M:%S").replace('NaT', 'Hora não registrada')
df_resumido["Empresa"] = col_empresa
df_resumido["Entregador"] = col_entregador
df_resumido["Pontualidade"] = col_horario_chegada.apply(lambda x: "chegou" if pd.notna(x) else "não chegou")
df_resumido["turno"] = df_resumido["Horario agendado"].apply(determinar_turno)

# Salvar a planilha tratada
output_path = os.path.join(pasta_planilhas, "Resumo_Tratado.xlsx")
df_resumido.to_excel(output_path, index=False)

# Abrir o arquivo automaticamente
os.startfile(output_path)
